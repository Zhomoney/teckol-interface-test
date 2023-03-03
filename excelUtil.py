import os

import openpyxl
from openpyxl import load_workbook, Workbook

import config


def FileIsExist(excelName):
    if not os.path.isfile(excelName):
        wbo0k = Workbook()
        wbo0k.active()
        wbo0k.save(config.testCasePath + '\\' + excelName)
        wbo0k.close()


def read_xlsx_excel(url, sheet_name):
    '''
    读取xlsx格式文件
    参数：
        url:文件路径
        sheet_name:表名
    返回：
        data:表格中的数据
    '''
    # 使用openpyxl加载指定路径的Excel文件并得到对应的workbook对象
    workbook = openpyxl.load_workbook(url)
    # 根据指定表名获取表格并得到对应的sheet对象
    sheet = workbook[sheet_name]
    # 定义列表存储表格数据
    data = []
    # 遍历表格的每一行
    for row in sheet.rows:
        # 定义表格存储每一行数据
        da = []
        # 从每一行中遍历每一个单元格
        for cell in row:
            # 将行数据存储到da列表
            da.append(cell.value)
        # 存储每一行数据
        data.append(da)
    # 返回数据
    return data


def write_xlsx_excel(url, sheet_name, two_dimensional_data):
    '''
    写入xlsx格式文件
    参数：
        url:文件路径
        sheet_name:表名
        two_dimensional_data：将要写入表格的数据（二维列表）
    '''
    # 创建工作簿对象
    workbook = load_workbook(url)
    # 创建工作表对象
    sheet = workbook.get_sheet_by_name(sheet_name)
    # 设置该工作表的名字

    # 遍历表格的每一行
    for i in range(0, len(two_dimensional_data)):
        # 遍历表格的每一列
        for j in range(0, len(two_dimensional_data[i])):
            # 写入数据（注意openpyxl的行和列是从1开始的，和我们平时的认知是一样的）
            sheet.cell(row=i + 1, column=j + 1, value=str(two_dimensional_data[i][j]))
    # 保存到指定位置
    workbook.save(url)
    print("写入成功")


def write_xlsx_excel_add(url, sheet_name, two_dimensional_data):
    '''
    追加写入xlsx格式文件
    参数：
        url:文件路径
        sheet_name:表名
        two_dimensional_data：将要写入表格的数据（二维列表）
    '''
    # 使用openpyxl加载指定路径的Excel文件并得到对应的workbook对象
    workbook = openpyxl.load_workbook(url)
    # 根据指定表名获取表格并得到对应的sheet对象
    sheet = workbook[sheet_name]
    for tdd in two_dimensional_data:
        sheet.append(tdd)
    # 保存到指定位置
    workbook.save(url)
    print("追加写入成功")
